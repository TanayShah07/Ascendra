from fastapi.responses import StreamingResponse

from io import BytesIO, StringIO

import csv

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from reportlab.lib.pagesizes import A4
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle
)
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet


# =========================================================
# BUILD USER DATA
# =========================================================

def build_user_data(user):

    return {

        "Profile": {

            "ID":
                user.id,

            "Full Name":
                user.full_name,

            "Email":
                user.email,

            "College":
                user.college,

            "Branch":
                user.branch,

            "Graduation Year":
                user.graduation_year,

            "Profile Photo":
                user.profile_photo,

            "Role":
                user.role,
        },

        "Professional Profiles": {

            "LinkedIn":
                user.linkedin,

            "GitHub":
                user.github,

            "Portfolio":
                user.portfolio,

            "LeetCode":
                user.leetcode,
        },

        "Placement": {

            "Dream Company":
                user.dream_company,

            "Target Role":
                user.target_role,

            "Preferred Domain":
                user.preferred_domain,

            "Placement Readiness":
                user.placement_readiness,
        },

        "Gamification": {

            "XP":
                user.xp,

            "Level":
                user.level,

            "Streak":
                user.streak,
        },

        "Resume": {

            "Resume URL":
                user.resume_url,

            "Resume Score":
                user.resume_score,

            "ATS Score":
                user.ats_score,
        },

        "Coding": {

            "Coding Problems Solved":
                user.coding_problems_solved,

            "Easy Solved":
                user.easy_solved,

            "Medium Solved":
                user.medium_solved,

            "Hard Solved":
                user.hard_solved,
        },

        "Progress": {

            "Interviews Completed":
                user.interview_completed,

            "GD Completed":
                user.gd_completed,

            "Roadmaps Generated":
                user.roadmap_generated,
        },

        "Account": {

            "Created At":
                str(user.created_at),
        }
    }


# =========================================================
# CSV EXPORT
# =========================================================

def generate_csv(data):

    output = StringIO()

    writer = csv.writer(output)

    writer.writerow([
        "Section",
        "Field",
        "Value"
    ])

    for section, fields in data.items():

        for field, value in fields.items():

            writer.writerow([
                section,
                field,
                "" if value is None else value
            ])

    output.seek(0)

    return StreamingResponse(

        iter([output.getvalue()]),

        media_type="text/csv",

        headers={
            "Content-Disposition":
                'attachment; filename="Ascendra_Data.csv"'
        }
    )


# =========================================================
# EXCEL EXPORT
# =========================================================

def generate_excel(data):

    workbook = Workbook()

    # Remove default sheet
    default_sheet = workbook.active

    workbook.remove(
        default_sheet
    )

    for section, fields in data.items():

        worksheet = workbook.create_sheet(
            title=section[:31]
        )

        worksheet.append([
            "Field",
            "Value"
        ])

        # Header styling
        for cell in worksheet[1]:

            cell.font = Font(
                bold=True
            )

            cell.fill = PatternFill(
                fill_type="solid",
                fgColor="2563EB"
            )

            cell.font = Font(
                bold=True,
                color="FFFFFF"
            )

            cell.alignment = Alignment(
                horizontal="center"
            )

        # Data
        for field, value in fields.items():

            worksheet.append([
                field,
                "" if value is None else value
            ])

        # Column widths
        worksheet.column_dimensions[
            "A"
        ].width = 32

        worksheet.column_dimensions[
            "B"
        ].width = 65

    output = BytesIO()

    workbook.save(output)

    output.seek(0)

    return StreamingResponse(

        output,

        media_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),

        headers={
            "Content-Disposition":
                'attachment; filename="Ascendra_Data.xlsx"'
        }
    )


# =========================================================
# PDF EXPORT
# =========================================================

def generate_pdf(data):

    output = BytesIO()

    document = SimpleDocTemplate(

        output,

        pagesize=A4,

        rightMargin=40,

        leftMargin=40,

        topMargin=40,

        bottomMargin=40,
    )

    styles = getSampleStyleSheet()

    title_style = styles["Title"]

    heading_style = styles["Heading2"]

    elements = []

    # Title
    elements.append(
        Paragraph(
            "Ascendra - Personal Data Export",
            title_style
        )
    )

    elements.append(
        Spacer(1, 20)
    )

    # Sections
    for section, fields in data.items():

        elements.append(
            Paragraph(
                section,
                heading_style
            )
        )

        elements.append(
            Spacer(1, 8)
        )

        table_data = [
            ["Field", "Value"]
        ]

        for field, value in fields.items():

            table_data.append([
                field,

                str(value)
                if value is not None
                else "Not provided"
            ])

        table = Table(
            table_data,
            colWidths=[
                160,
                320
            ]
        )

        table.setStyle(
            TableStyle([

                (
                    "BACKGROUND",
                    (0, 0),
                    (-1, 0),
                    colors.HexColor("#2563EB")
                ),

                (
                    "TEXTCOLOR",
                    (0, 0),
                    (-1, 0),
                    colors.white
                ),

                (
                    "FONTNAME",
                    (0, 0),
                    (-1, 0),
                    "Helvetica-Bold"
                ),

                (
                    "FONTNAME",
                    (0, 1),
                    (-1, -1),
                    "Helvetica"
                ),

                (
                    "GRID",
                    (0, 0),
                    (-1, -1),
                    0.5,
                    colors.grey
                ),

                (
                    "VALIGN",
                    (0, 0),
                    (-1, -1),
                    "TOP"
                ),

                (
                    "LEFTPADDING",
                    (0, 0),
                    (-1, -1),
                    8
                ),

                (
                    "RIGHTPADDING",
                    (0, 0),
                    (-1, -1),
                    8
                ),

                (
                    "TOPPADDING",
                    (0, 0),
                    (-1, -1),
                    6
                ),

                (
                    "BOTTOMPADDING",
                    (0, 0),
                    (-1, -1),
                    6
                ),

            ])
        )

        elements.append(table)

        elements.append(
            Spacer(1, 20)
        )

    document.build(elements)

    output.seek(0)

    return StreamingResponse(

        output,

        media_type="application/pdf",

        headers={
            "Content-Disposition":
                'attachment; filename="Ascendra_Data.pdf"'
        }
    )


# =========================================================
# MAIN EXPORT FUNCTION
# =========================================================

def generate_user_export(
    user,
    format
):

    data = build_user_data(user)

    # -----------------------------------------------------
    # PDF
    # -----------------------------------------------------

    if format == "pdf":

        return generate_pdf(
            data
        )

    # -----------------------------------------------------
    # EXCEL
    # -----------------------------------------------------

    if format == "xlsx":

        return generate_excel(
            data
        )

    # -----------------------------------------------------
    # CSV
    # -----------------------------------------------------

    if format == "csv":

        return generate_csv(
            data
        )

    # -----------------------------------------------------
    # INVALID FORMAT
    # -----------------------------------------------------

    raise ValueError(
        "Unsupported export format. "
        "Use pdf, xlsx, or csv."
    )